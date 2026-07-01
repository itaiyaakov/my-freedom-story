import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
import json
import os
import io
import base64
from datetime import date, timedelta, datetime

# ─── קבועים ─────────────────────────────────────────────────────────────────
DATA_FILE = "calorie_data.json"

ACTIVITY_FACTORS = {
    "יושבני (כמעט ללא פעילות)": 1.2,
    "פעיל קלות (1–3 אימונים בשבוע)": 1.375,
    "פעיל בינוני (3–5 אימונים בשבוע)": 1.55,
    "פעיל מאוד (6–7 אימונים בשבוע)": 1.725,
    "ספורטאי (עבודה פיזית + אימונים)": 1.9,
}

COMMON_FOODS = {
    "ארוחת בוקר": [
        ("ביצה קשה", 78), ("ביצה מקושקשת (2)", 180), ("יוגורט 0%", 60),
        ("גבינה לבנה 5% (100g)", 105), ("לחם מחיטה מלאה (פרוסה)", 80),
        ("שיבולת שועל (50g יבש)", 177), ("בננה", 90), ("תפוח", 80),
    ],
    "ארוחת צהריים": [
        ("חזה עוף (150g)", 248), ("סלמון (150g)", 280), ("אורז לבן (מנה)", 206),
        ("פסטה (מנה)", 220), ("סלט ירקות", 50), ("עדשים מבושלות (100g)", 116),
        ("טוסט עם אבוקדו", 250), ("פלאפל (4 יח')", 280),
    ],
    "ארוחת ערב": [
        ("מרק עוף", 80), ("קוביות טופו (100g)", 144), ("בשר בקר רזה (150g)", 310),
        ("ירקות מוקפצים", 120), ("קינואה (100g מבושל)", 120), ("שניצל (150g)", 350),
    ],
    "חטיפים": [
        ("שקדים (30g)", 173), ("אגוז מלך (30g)", 196), ("חומוס (4 כפות)", 166),
        ("פרי (בממוצע)", 80), ("שוקולד מריר (25g)", 142), ("גרנולה (50g)", 220),
    ],
}

COMMON_DRINKS = [
    ("מים (250ml)", 0), ("קפה שחור", 5), ("קפה עם חלב", 35),
    ("תה (ללא סוכר)", 2), ("מיץ תפוזים (200ml)", 90), ("מיץ ירקות (200ml)", 60),
    ("חלב 1% (200ml)", 82), ("חלב שקדים (200ml)", 30), ("שייק חלבון", 150),
    ("קולה (330ml)", 140), ("קולה זירו (330ml)", 2), ("בירה (330ml)", 145),
    ("יין אדום (125ml)", 108), ("מים מוגזים", 0),
]

COMMON_ACTIVITIES = [
    ("הליכה קלה (30 דק')", 120), ("הליכה מהירה (30 דק')", 180),
    ("ריצה קלה (30 דק')", 280), ("ריצה מהירה (30 דק')", 380),
    ("אופניים (30 דק')", 250), ("שחייה (30 דק')", 300),
    ("אימון כוח (45 דק')", 220), ("יוגה (45 דק')", 150),
    ("ספינינג (45 דק')", 400), ("כדורסל (30 דק')", 260),
    ("כדורגל (30 דק')", 270), ("טניס (30 דק')", 230),
    ("קפיצות חבל (15 דק')", 200), ("מדרגות (15 דק')", 130),
]

# ─── טעינה ושמירה ────────────────────────────────────────────────────────────
def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {
        "profile": {},
        "diary": {},
        "weight_log": {},
    }


def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ─── חישובים ─────────────────────────────────────────────────────────────────
def calc_bmr(weight, height, age, gender):
    if gender == "זכר":
        return 10 * weight + 6.25 * height - 5 * age + 5
    return 10 * weight + 6.25 * height - 5 * age - 161


def calc_tdee(bmr, activity_level):
    return bmr * ACTIVITY_FACTORS[activity_level]


def days_to_goal(current_weight, goal_weight, daily_deficit):
    weight_diff = current_weight - goal_weight
    if weight_diff <= 0 or daily_deficit <= 0:
        return None
    calories_to_burn = weight_diff * 7700
    return int(calories_to_burn / daily_deficit)


# ─── יצוא אקסל ───────────────────────────────────────────────────────────────
def generate_excel(data):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1E3A8A")

    def style_header(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ── גיליון 1: יומן יומי ──
    ws1 = wb.active
    ws1.title = "יומן יומי"
    ws1.sheet_view.rightToLeft = True
    headers = ["תאריך", "סוג", "פריט", "קלוריות", "כמות"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        style_header(cell)
    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["C"].width = 28

    row = 2
    for date_key in sorted(data.get("diary", {}).keys()):
        day = data["diary"][date_key]
        for item in day.get("food", []):
            ws1.cell(row=row, column=1, value=date_key)
            ws1.cell(row=row, column=2, value="מזון")
            ws1.cell(row=row, column=3, value=item["name"])
            ws1.cell(row=row, column=4, value=item["cal"])
            ws1.cell(row=row, column=5, value=item.get("qty", 1))
            row += 1
        for item in day.get("drinks", []):
            ws1.cell(row=row, column=1, value=date_key)
            ws1.cell(row=row, column=2, value="שתייה")
            ws1.cell(row=row, column=3, value=item["name"])
            ws1.cell(row=row, column=4, value=item["cal"])
            ws1.cell(row=row, column=5, value=item.get("qty", 1))
            row += 1
        for item in day.get("activities", []):
            ws1.cell(row=row, column=1, value=date_key)
            ws1.cell(row=row, column=2, value="פעילות")
            ws1.cell(row=row, column=3, value=item["name"])
            ws1.cell(row=row, column=4, value=-item["cal"])
            ws1.cell(row=row, column=5, value=1)
            row += 1

    # ── גיליון 2: מעקב משקל ──
    ws2 = wb.create_sheet("מעקב משקל")
    ws2.sheet_view.rightToLeft = True
    for col, h in enumerate(["תאריך", "משקל (ק\"ג)"], 1):
        style_header(ws2.cell(row=1, column=col, value=h))
    ws2.column_dimensions["A"].width = 14
    for i, (dk, w) in enumerate(sorted(data.get("weight_log", {}).items()), 2):
        ws2.cell(row=i, column=1, value=dk)
        ws2.cell(row=i, column=2, value=w)

    # ── גיליון 3: פרופיל ──
    ws3 = wb.create_sheet("פרופיל")
    ws3.sheet_view.rightToLeft = True
    profile = data.get("profile", {})
    fields = [
        ("שם", "name"), ("גיל", "age"), ("גובה (ס\"מ)", "height"),
        ("מגדר", "gender"), ("משקל נוכחי (ק\"ג)", "weight"),
        ("משקל יעד (ק\"ג)", "goal_weight"), ("BMR", "bmr"), ("TDEE", "tdee"),
        ("גרעון יומי מבוקש", "deficit_target"),
    ]
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 18
    for i, (label, key) in enumerate(fields, 1):
        cell = ws3.cell(row=i, column=1, value=label)
        cell.font = Font(bold=True)
        ws3.cell(row=i, column=2, value=profile.get(key, ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── ביקורת AI ───────────────────────────────────────────────────────────────
def ai_meal_review(day_data, profile, api_key):
    import anthropic

    food_lines  = [f"• {i['name']} ({i['cal']} קל)" for i in day_data.get("food", [])]
    drink_lines = [f"• {i['name']} ({i['cal']} קל)" for i in day_data.get("drinks", [])]
    act_lines   = [f"• {i['name']} (שרפה {i['cal']} קל)" for i in day_data.get("activities", [])]

    total_in  = sum(i["cal"] for i in day_data.get("food", [])) + \
                sum(i["cal"] for i in day_data.get("drinks", []))
    total_burn= sum(i["cal"] for i in day_data.get("activities", []))
    tdee      = profile.get("tdee", 2000)
    target    = tdee - profile.get("deficit_target", 500)

    prompt = f"""אתה דיאטנית מומחית ומעודדת. בקר את הרישום התזונתי הבא בעברית בלבד.

**מזון שנאכל היום:**
{chr(10).join(food_lines) if food_lines else "לא הוזן מזון"}

**שתייה:**
{chr(10).join(drink_lines) if drink_lines else "לא הוזן"}

**פעילות גופנית:**
{chr(10).join(act_lines) if act_lines else "לא הוזן"}

**סה"כ קלוריות שנצרכו:** {total_in} קל
**שריפה בפעילות:** {total_burn} קל
**נטו:** {total_in - total_burn} קל
**יעד יומי:** {target:.0f} קל

כתוב ביקורת קצרה ומועילה הכוללת בדיוק את 4 הסעיפים הבאים:

✅ **מה עשית טוב** – ציין 2-3 דברים חיוביים ספציפיים
⚠️ **מה כדאי לשפר** – ציין 1-2 בעיות תזונתיות
🔄 **חלופות בריאות יותר** – לכל פריט לא בריא שהוזן, הצע חלופה ספציפית (למשל: "במקום ציפס — נסה גזר עם חומוס")
💡 **טיפ אחד להמשך** – עצה מעשית אחת לשיפור מחר

היה ספציפי לפריטים שהוזנו. אל תחזור על הרשימה. כתוב בשפה עברית חמה ומעודדת."""

    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-opus-4-8",
        max_tokens=1024,
        messages=[{"role": "user", "content": prompt}]
    )
    return message.content[0].text


# ─── זיהוי מזון מתמונה ───────────────────────────────────────────────────────
def analyze_food_photo(image_bytes, media_type, api_key):
    import anthropic

    client = anthropic.Anthropic(api_key=api_key)
    image_data = base64.standard_b64encode(image_bytes).decode("utf-8")

    message = client.messages.create(
        model="claude-opus-4-8",
        max_tokens=1024,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {"type": "base64", "media_type": media_type, "data": image_data},
                },
                {
                    "type": "text",
                    "text": (
                        "זהה את כל פריטי המזון בתמונה ואמוד את הקלוריות של כל אחד.\n"
                        "החזר JSON בלבד (ללא טקסט נוסף), בפורמט:\n"
                        "{\"items\": [{\"name\": \"שם בעברית\", \"cal\": 250, \"qty\": 1}, ...]}\n"
                        "היה ספציפי לגבי הכמות/גודל המנה. אם אינך בטוח, אמוד בצד הגבוה."
                    ),
                },
            ],
        }],
    )

    text = message.content[0].text.strip()
    if "```" in text:
        text = text.split("```")[1]
        if text.startswith("json"):
            text = text[4:]
    return json.loads(text.strip())


# ─── CSS ─────────────────────────────────────────────────────────────────────
def apply_css():
    st.markdown("""
    <style>
    .stApp { direction: rtl; text-align: right; background-color: #f8fafc; }
    h1, h2, h3, h4 { text-align: right; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; flex-direction: row-reverse; }
    .stTabs [data-baseweb="tab"] { direction: rtl; }

    .metric-card {
        background: white;
        border-radius: 16px;
        padding: 20px;
        text-align: center;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08);
        margin-bottom: 12px;
    }
    .metric-value { font-size: 32px; font-weight: 800; margin: 4px 0; }
    .metric-label { font-size: 13px; color: #64748b; margin: 0; }

    .deficit-card {
        border-radius: 16px;
        padding: 24px;
        text-align: center;
        margin: 12px 0;
        font-size: 20px;
        font-weight: 700;
    }
    .entry-row {
        background: white;
        border-radius: 10px;
        padding: 10px 16px;
        margin: 6px 0;
        display: flex;
        justify-content: space-between;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    }
    .section-header {
        font-size: 18px;
        font-weight: 700;
        color: #1e293b;
        margin: 20px 0 10px;
        padding-bottom: 6px;
        border-bottom: 2px solid #e2e8f0;
    }
    .ai-review-box {
        background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
        border-radius: 16px;
        padding: 24px;
        border: 1px solid #bae6fd;
        margin-top: 16px;
        line-height: 1.8;
    }
    div[data-testid="stNumberInput"] { direction: ltr; }
    </style>
    """, unsafe_allow_html=True)


# ─── עמוד: פרופיל אישי ───────────────────────────────────────────────────────
def page_profile(data):
    st.header("פרופיל אישי")
    p = data.get("profile", {})

    with st.form("profile_form"):
        c1, c2 = st.columns(2)
        with c1:
            name   = st.text_input("שם", value=p.get("name", ""))
            age    = st.number_input("גיל", 10, 100, int(p.get("age", 30)))
            height = st.number_input("גובה (ס\"מ)", 100, 220, int(p.get("height", 170)))
        with c2:
            gender  = st.radio("מגדר", ["זכר", "נקבה"], index=0 if p.get("gender", "זכר") == "זכר" else 1)
            weight  = st.number_input("משקל נוכחי (ק\"ג)", 30.0, 250.0, float(p.get("weight", 75.0)), step=0.5)
            g_weight = st.number_input("משקל יעד (ק\"ג)", 30.0, 250.0, float(p.get("goal_weight", 70.0)), step=0.5)

        activity = st.selectbox("רמת פעילות", list(ACTIVITY_FACTORS.keys()),
                                index=list(ACTIVITY_FACTORS.keys()).index(p.get("activity", list(ACTIVITY_FACTORS.keys())[1])))

        deficit_target = st.slider("גרעון יומי מבוקש (קלוריות)", 0, 1200,
                                   int(p.get("deficit_target", 500)), step=50,
                                   help="500 = ירידה של כ-0.5 ק\"ג בשבוע | 1000 = כ-1 ק\"ג בשבוע")

        saved = st.form_submit_button("שמור פרופיל", use_container_width=True)

    if saved:
        bmr  = calc_bmr(weight, height, age, gender)
        tdee = calc_tdee(bmr, activity)
        data["profile"] = {
            "name": name, "age": age, "height": height, "gender": gender,
            "weight": weight, "goal_weight": g_weight,
            "activity": activity, "deficit_target": deficit_target,
            "bmr": round(bmr), "tdee": round(tdee),
        }
        today = str(date.today())
        if today not in data["weight_log"]:
            data["weight_log"][today] = weight
        save_data(data)
        st.success("הפרופיל נשמר!")
        st.rerun()

    if data.get("profile"):
        p = data["profile"]
        bmr  = p.get("bmr", 0)
        tdee = p.get("tdee", 0)
        target_cal = tdee - p.get("deficit_target", 500)
        st.markdown("---")
        st.subheader("הנתונים שלך")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(f"""<div class='metric-card'>
                <p class='metric-label'>BMR (חילוף חומרים בסיסי)</p>
                <p class='metric-value' style='color:#3b82f6'>{bmr:,}</p>
                <p class='metric-label'>קלוריות</p></div>""", unsafe_allow_html=True)
        with c2:
            st.markdown(f"""<div class='metric-card'>
                <p class='metric-label'>TDEE (צריכה יומית)</p>
                <p class='metric-value' style='color:#8b5cf6'>{tdee:,}</p>
                <p class='metric-label'>קלוריות</p></div>""", unsafe_allow_html=True)
        with c3:
            st.markdown(f"""<div class='metric-card'>
                <p class='metric-label'>יעד קלורי יומי</p>
                <p class='metric-value' style='color:#10b981'>{target_cal:,.0f}</p>
                <p class='metric-label'>קלוריות</p></div>""", unsafe_allow_html=True)
        with c4:
            days = days_to_goal(p.get("weight", 75), p.get("goal_weight", 70), p.get("deficit_target", 500))
            goal_date = (date.today() + timedelta(days=days)).strftime("%d/%m/%Y") if days else "—"
            st.markdown(f"""<div class='metric-card'>
                <p class='metric-label'>תאריך יעד משוער</p>
                <p class='metric-value' style='color:#f59e0b; font-size:20px'>{goal_date}</p>
                <p class='metric-label'>{days if days else '—'} ימים</p></div>""", unsafe_allow_html=True)


# ─── עמוד: יומן יומי ────────────────────────────────────────────────────────
def page_diary(data):
    st.header("יומן יומי")

    today_str = str(date.today())
    selected_date = st.date_input("תאריך", value=date.today(), max_value=date.today())
    day_key = str(selected_date)

    if day_key not in data["diary"]:
        data["diary"][day_key] = {"food": [], "drinks": [], "activities": []}

    day = data["diary"][day_key]

    # ─── הוספת מזון ──────────────────────────────────────────────────────────
    st.markdown("<div class='section-header'>🍽️ ארוחות ואוכל</div>", unsafe_allow_html=True)
    tab_a, tab_b, tab_photo = st.tabs(["רשימה מהירה", "הזנה ידנית", "📷 צילום"])

    with tab_photo:
        api_key_photo = st.session_state.get("anthropic_api_key", "")
        if not api_key_photo:
            st.info("הזן מפתח API בסרגל הצד כדי לזהות מזון מתמונה")
        else:
            st.caption("צלם את הצלחת שלך — Claude יזהה את המאכלים ויאמוד קלוריות")
            img_source = st.radio("מקור תמונה", ["מצלמה", "גלריה"], horizontal=True, key="img_source")

            uploaded = None
            if img_source == "מצלמה":
                uploaded = st.camera_input("צלם את הארוחה", key="food_camera")
            else:
                uploaded = st.file_uploader("בחר תמונה מהגלריה", type=["jpg", "jpeg", "png", "webp"], key="food_upload")

            if uploaded and st.button("🔍 זהה מזון", key="analyze_photo", use_container_width=True, type="primary"):
                with st.spinner("Claude מזהה את המאכלים..."):
                    try:
                        img_bytes = uploaded.getvalue()
                        ext = uploaded.name.split(".")[-1].lower() if hasattr(uploaded, "name") else "jpeg"
                        mime_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp"}
                        media_type = mime_map.get(ext, "image/jpeg")
                        result = analyze_food_photo(img_bytes, media_type, api_key_photo)
                        st.session_state["photo_items"] = result.get("items", [])
                    except Exception as e:
                        st.error(f"שגיאה בזיהוי: {e}")

            if st.session_state.get("photo_items"):
                st.markdown("**פריטים שזוהו — בחר מה להוסיף:**")
                selected = []
                for idx, item in enumerate(st.session_state["photo_items"]):
                    col_chk, col_name, col_cal = st.columns([0.5, 3, 1.5])
                    with col_chk:
                        checked = st.checkbox("", value=True, key=f"photo_chk_{idx}")
                    with col_name:
                        item["name"] = st.text_input("", value=item["name"], key=f"photo_name_{idx}", label_visibility="collapsed")
                    with col_cal:
                        item["cal"] = st.number_input("", value=int(item["cal"]), min_value=0, max_value=5000, key=f"photo_cal_{idx}", label_visibility="collapsed")
                    if checked:
                        selected.append(item)

                if st.button("✅ הוסף פריטים נבחרים", key="add_photo_items", use_container_width=True):
                    for item in selected:
                        day["food"].append({"name": item["name"], "cal": item["cal"], "qty": item.get("qty", 1)})
                    save_data(data)
                    st.session_state["photo_items"] = []
                    st.rerun()

    with tab_a:
        category = st.selectbox("קטגוריה", list(COMMON_FOODS.keys()), key="food_cat")
        food_options = {f"{name} ({cal} קל)": (name, cal) for name, cal in COMMON_FOODS[category]}
        chosen = st.selectbox("בחר מזון", list(food_options.keys()), key="food_select")
        qty = st.number_input("כמות (מנות)", 0.5, 10.0, 1.0, step=0.5, key="food_qty")
        if st.button("הוסף לארוחה", key="add_food_quick"):
            name_, cal_ = food_options[chosen]
            day["food"].append({"name": name_, "cal": round(cal_ * qty), "qty": qty})
            save_data(data)
            st.rerun()

    with tab_b:
        c1, c2 = st.columns([3, 1])
        with c1: custom_food = st.text_input("שם המאכל", key="custom_food_name")
        with c2: custom_cal  = st.number_input("קלוריות", 0, 5000, 200, key="custom_food_cal")
        if st.button("הוסף", key="add_food_manual"):
            if custom_food:
                day["food"].append({"name": custom_food, "cal": custom_cal, "qty": 1})
                save_data(data)
                st.rerun()

    if day["food"]:
        for i, item in enumerate(day["food"]):
            c1, c2 = st.columns([5, 1])
            with c1:
                qty_str = f" ×{item['qty']}" if item.get("qty", 1) != 1 else ""
                st.write(f"🍴 {item['name']}{qty_str} — **{item['cal']} קל**")
            with c2:
                if st.button("✕", key=f"del_food_{i}"):
                    day["food"].pop(i)
                    save_data(data)
                    st.rerun()

    # ─── שתייה ───────────────────────────────────────────────────────────────
    st.markdown("<div class='section-header'>💧 שתייה</div>", unsafe_allow_html=True)
    tab_c, tab_d = st.tabs(["רשימה מהירה", "הזנה ידנית"])

    with tab_c:
        drink_options = {f"{name} ({cal} קל)": (name, cal) for name, cal in COMMON_DRINKS}
        d_chosen = st.selectbox("בחר שתייה", list(drink_options.keys()), key="drink_select")
        d_qty = st.number_input("כמות (כוסות/יחידות)", 1, 20, 1, key="drink_qty")
        if st.button("הוסף שתייה", key="add_drink_quick"):
            d_name, d_cal = drink_options[d_chosen]
            day["drinks"].append({"name": d_name, "cal": d_cal * d_qty, "qty": d_qty})
            save_data(data)
            st.rerun()

    with tab_d:
        c1, c2 = st.columns([3, 1])
        with c1: custom_drink = st.text_input("שם המשקה", key="custom_drink_name")
        with c2: custom_dcal  = st.number_input("קלוריות", 0, 2000, 0, key="custom_drink_cal")
        if st.button("הוסף", key="add_drink_manual"):
            if custom_drink:
                day["drinks"].append({"name": custom_drink, "cal": custom_dcal, "qty": 1})
                save_data(data)
                st.rerun()

    if day["drinks"]:
        for i, item in enumerate(day["drinks"]):
            c1, c2 = st.columns([5, 1])
            with c1:
                qty_str = f" ×{item['qty']}" if item.get("qty", 1) != 1 else ""
                st.write(f"💧 {item['name']}{qty_str} — **{item['cal']} קל**")
            with c2:
                if st.button("✕", key=f"del_drink_{i}"):
                    day["drinks"].pop(i)
                    save_data(data)
                    st.rerun()

    # ─── פעילות גופנית ────────────────────────────────────────────────────────
    st.markdown("<div class='section-header'>🏃 פעילות גופנית</div>", unsafe_allow_html=True)
    tab_e, tab_f = st.tabs(["רשימה מהירה", "הזנה ידנית"])

    with tab_e:
        act_options = {f"{name} ({cal} קל)": (name, cal) for name, cal in COMMON_ACTIVITIES}
        a_chosen = st.selectbox("בחר פעילות", list(act_options.keys()), key="act_select")
        if st.button("הוסף פעילות", key="add_act_quick"):
            a_name, a_cal = act_options[a_chosen]
            day["activities"].append({"name": a_name, "cal": a_cal})
            save_data(data)
            st.rerun()

    with tab_f:
        c1, c2 = st.columns([3, 1])
        with c1: custom_act  = st.text_input("שם הפעילות", key="custom_act_name")
        with c2: custom_acal = st.number_input("קלוריות שנשרפו", 0, 3000, 200, key="custom_act_cal")
        if st.button("הוסף", key="add_act_manual"):
            if custom_act:
                day["activities"].append({"name": custom_act, "cal": custom_acal})
                save_data(data)
                st.rerun()

    if day["activities"]:
        for i, item in enumerate(day["activities"]):
            c1, c2 = st.columns([5, 1])
            with c1: st.write(f"🔥 {item['name']} — **{item['cal']} קל שנשרפו**")
            with c2:
                if st.button("✕", key=f"del_act_{i}"):
                    day["activities"].pop(i)
                    save_data(data)
                    st.rerun()

    # ─── סיכום יומי ──────────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("סיכום יומי")
    total_in   = sum(i["cal"] for i in day["food"]) + sum(i["cal"] for i in day["drinks"])
    total_burn = sum(i["cal"] for i in day["activities"])
    net_cal    = total_in - total_burn

    profile = data.get("profile", {})
    tdee    = profile.get("tdee", 2000)
    target  = tdee - profile.get("deficit_target", 500)
    balance = target - net_cal

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f"""<div class='metric-card'>
            <p class='metric-label'>צריכה</p>
            <p class='metric-value' style='color:#ef4444'>{total_in:,}</p>
            <p class='metric-label'>קלוריות</p></div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class='metric-card'>
            <p class='metric-label'>שריפה בפעילות</p>
            <p class='metric-value' style='color:#10b981'>{total_burn:,}</p>
            <p class='metric-label'>קלוריות</p></div>""", unsafe_allow_html=True)
    with c3:
        st.markdown(f"""<div class='metric-card'>
            <p class='metric-label'>נטו</p>
            <p class='metric-value' style='color:#6366f1'>{net_cal:,}</p>
            <p class='metric-label'>קלוריות</p></div>""", unsafe_allow_html=True)

    if profile:
        if balance > 0:
            color, emoji, msg = "#10b981", "✅", f"גרעון של {balance:,} קל — מעולה!"
        elif balance == 0:
            color, emoji, msg = "#f59e0b", "⚖️", "בדיוק על היעד!"
        else:
            color, emoji, msg = "#ef4444", "⚠️", f"עודף של {abs(balance):,} קל"

        st.markdown(f"""<div class='deficit-card' style='background:{color}22; border: 2px solid {color};'>
            {emoji} {msg}<br>
            <small style='font-weight:400;color:#475569'>יעד יומי: {target:,.0f} קל | TDEE: {tdee:,} קל</small>
        </div>""", unsafe_allow_html=True)

        pct = min(1.0, net_cal / target) if target > 0 else 0
        st.progress(pct, text=f"נצרכו {pct*100:.0f}% מהיעד היומי")

    # ─── גרף עוגה יומי ───────────────────────────────────────────────────────
    if day["food"] or day["drinks"]:
        all_items = [(i["name"], i["cal"]) for i in day["food"]] + \
                    [(i["name"], i["cal"]) for i in day["drinks"]]
        names = [x[0] for x in all_items]
        cals  = [x[1] for x in all_items]
        fig = px.pie(values=cals, names=names,
                     title="פילוח קלוריות יומי", hole=0.4)
        fig.update_layout(font=dict(family="Arial"), title_x=0.5)
        st.plotly_chart(fig, use_container_width=True)

    # ─── ביקורת AI ───────────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("🤖 ביקורת תזונה חכמה")

    api_key = st.session_state.get("anthropic_api_key", "")
    has_entries = bool(day["food"] or day["drinks"])

    if not api_key:
        st.info("הזן מפתח API של Anthropic בסרגל הצד כדי לקבל ביקורת תזונה מ-AI")
    elif not has_entries:
        st.info("הזן מאכלים או שתייה כדי לקבל ביקורת תזונה")
    else:
        if st.button("🔍 בקש ביקורת תזונה מה-AI", use_container_width=True, type="primary"):
            with st.spinner("הדיאטנית החכמה מנתחת את הארוחות שלך..."):
                try:
                    review = ai_meal_review(day, profile, api_key)
                    st.session_state["last_review"] = review
                    st.session_state["last_review_date"] = day_key
                except Exception as e:
                    st.error(f"שגיאה בקריאה ל-AI: {e}")

        if st.session_state.get("last_review") and st.session_state.get("last_review_date") == day_key:
            st.markdown(
                f"<div class='ai-review-box'>{st.session_state['last_review'].replace(chr(10), '<br>')}</div>",
                unsafe_allow_html=True
            )


# ─── עמוד: דשבורד ────────────────────────────────────────────────────────────
def page_dashboard(data):
    st.header("דשבורד מעקב")

    profile = data.get("profile", {})
    if not profile:
        st.info("מלא קודם את הפרופיל האישי שלך")
        return

    # ─── יצוא אקסל ───────────────────────────────────────────────────────────
    col_export, _ = st.columns([1, 3])
    with col_export:
        try:
            excel_bytes = generate_excel(data)
            filename = f"calorie_tracker_{date.today().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                label="📥 יצא לאקסל",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except ImportError:
            st.warning("נדרש: pip install openpyxl")

    tdee   = profile.get("tdee", 2000)
    target = tdee - profile.get("deficit_target", 500)
    goal_w = profile.get("goal_weight", 70)
    curr_w = profile.get("weight", 75)

    # ─── ריכוז 7 ימים אחרונים ─────────────────────────────────────────────────
    dates, cals_in, cals_burn, cals_net = [], [], [], []
    for i in range(6, -1, -1):
        d = date.today() - timedelta(days=i)
        dk = str(d)
        day = data["diary"].get(dk, {"food": [], "drinks": [], "activities": []})
        tin  = sum(x["cal"] for x in day["food"]) + sum(x["cal"] for x in day["drinks"])
        tburn= sum(x["cal"] for x in day["activities"])
        dates.append(d.strftime("%d/%m"))
        cals_in.append(tin)
        cals_burn.append(tburn)
        cals_net.append(tin - tburn)

    fig = go.Figure()
    fig.add_bar(name="צריכה", x=dates, y=cals_in, marker_color="#ef4444")
    fig.add_bar(name="שריפה", x=dates, y=cals_burn, marker_color="#10b981")
    fig.add_scatter(name="יעד יומי", x=dates, y=[target]*7,
                    mode="lines", line=dict(color="#6366f1", dash="dash", width=2))
    fig.update_layout(
        title="קלוריות – 7 ימים אחרונים", barmode="group",
        xaxis_title="תאריך", yaxis_title="קלוריות",
        font=dict(family="Arial"), title_x=0.5,
        legend=dict(orientation="h", y=-0.2)
    )
    st.plotly_chart(fig, use_container_width=True)

    deficits = [target - n for n in cals_net]
    colors_d = ["#10b981" if d >= 0 else "#ef4444" for d in deficits]
    fig2 = go.Figure(go.Bar(
        x=dates, y=deficits,
        marker_color=colors_d,
        text=[f"{d:+.0f}" for d in deficits],
        textposition="outside",
    ))
    fig2.add_hline(y=0, line_dash="dot", line_color="#94a3b8")
    fig2.update_layout(
        title="גרעון / עודף קלורי יומי", title_x=0.5,
        yaxis_title="קלוריות (חיובי = גרעון)", font=dict(family="Arial")
    )
    st.plotly_chart(fig2, use_container_width=True)

    # ─── מעקב משקל ────────────────────────────────────────────────────────────
    st.subheader("מעקב משקל")
    col1, col2 = st.columns([2, 1])
    with col2:
        today_str = str(date.today())
        new_weight = st.number_input(
            "רשום משקל היום (ק\"ג)",
            30.0, 250.0,
            float(data["weight_log"].get(today_str, curr_w)),
            step=0.1, key="weight_log_input"
        )
        if st.button("שמור משקל", key="save_weight"):
            data["weight_log"][today_str] = new_weight
            data["profile"]["weight"] = new_weight
            save_data(data)
            st.success("נשמר!")
            st.rerun()

    with col1:
        wlog = data.get("weight_log", {})
        if len(wlog) >= 2:
            sorted_dates = sorted(wlog.keys())
            w_dates = [datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m") for d in sorted_dates]
            w_vals  = [wlog[d] for d in sorted_dates]
            fig3 = go.Figure()
            fig3.add_scatter(x=w_dates, y=w_vals, mode="lines+markers",
                             line=dict(color="#6366f1", width=3),
                             marker=dict(size=8, color="#6366f1"),
                             name="משקל")
            fig3.add_hline(y=goal_w, line_dash="dash",
                           line_color="#10b981", annotation_text=f"יעד {goal_w} ק\"ג")
            fig3.update_layout(
                title="היסטוריית משקל", title_x=0.5,
                yaxis_title="ק\"ג", font=dict(family="Arial"),
                yaxis=dict(range=[min(min(w_vals), goal_w) - 2, max(max(w_vals), curr_w) + 2])
            )
            st.plotly_chart(fig3, use_container_width=True)
        else:
            st.info("הוסף לפחות 2 מדידות משקל כדי לראות גרף")

    # ─── סטטיסטיקות שבועיות ──────────────────────────────────────────────────
    st.subheader("סיכום שבועי")
    avg_in   = sum(cals_in) / 7
    avg_burn = sum(cals_burn) / 7
    avg_net  = sum(cals_net) / 7
    total_def= sum(deficits)
    est_loss = total_def / 7700

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f"""<div class='metric-card'>
            <p class='metric-label'>ממוצע צריכה יומית</p>
            <p class='metric-value' style='color:#ef4444'>{avg_in:,.0f}</p>
            <p class='metric-label'>קל/יום</p></div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class='metric-card'>
            <p class='metric-label'>ממוצע שריפה יומית</p>
            <p class='metric-value' style='color:#10b981'>{avg_burn:,.0f}</p>
            <p class='metric-label'>קל/יום</p></div>""", unsafe_allow_html=True)
    with c3:
        st.markdown(f"""<div class='metric-card'>
            <p class='metric-label'>גרעון שבועי כולל</p>
            <p class='metric-value' style='color:#6366f1'>{total_def:+,.0f}</p>
            <p class='metric-label'>קלוריות</p></div>""", unsafe_allow_html=True)
    with c4:
        color = "#10b981" if est_loss >= 0 else "#ef4444"
        label = "ירידה משוערת" if est_loss >= 0 else "עלייה משוערת"
        st.markdown(f"""<div class='metric-card'>
            <p class='metric-label'>{label}</p>
            <p class='metric-value' style='color:{color}'>{abs(est_loss):.2f}</p>
            <p class='metric-label'>ק\"ג השבוע</p></div>""", unsafe_allow_html=True)

    wlog_sorted = sorted(data.get("weight_log", {}).items())
    if wlog_sorted:
        start_w  = wlog_sorted[0][1]
        latest_w = wlog_sorted[-1][1]
        if start_w > goal_w:
            total_to_lose = start_w - goal_w
            lost_so_far   = start_w - latest_w
            pct_done      = max(0, min(1, lost_so_far / total_to_lose))
            st.subheader("התקדמות לעבר היעד")
            st.progress(pct_done, text=f"ירדת {lost_so_far:.1f} ק\"ג מתוך {total_to_lose:.1f} ק\"ג ({pct_done*100:.1f}%)")

            days_left = days_to_goal(latest_w, goal_w, profile.get("deficit_target", 500))
            if days_left:
                eta = date.today() + timedelta(days=days_left)
                st.info(f"בקצב הנוכחי תגיע ליעד בערך ב-{eta.strftime('%d/%m/%Y')} (עוד {days_left} ימים)")


# ─── ראשי ─────────────────────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="מעקב קלוריות",
        page_icon="🥗",
        layout="wide",
    )
    apply_css()

    # ─── מפתח API בסרגל הצד ──────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### 🤖 הגדרות AI")
        # נסה לטעון ממשתנה סביבה / secrets
        default_key = ""
        try:
            default_key = st.secrets.get("ANTHROPIC_API_KEY", "")
        except Exception:
            default_key = os.environ.get("ANTHROPIC_API_KEY", "")

        if default_key:
            st.session_state["anthropic_api_key"] = default_key
            st.success("מפתח API נטען אוטומטית ✓")
        else:
            entered_key = st.text_input(
                "מפתח Anthropic API",
                value=st.session_state.get("anthropic_api_key", ""),
                type="password",
                placeholder="sk-ant-...",
                help="נדרש לביקורת תזונה חכמה. ניתן לקבל ב-console.anthropic.com",
            )
            if entered_key:
                st.session_state["anthropic_api_key"] = entered_key

        st.markdown("---")
        st.markdown("### 📱 קיצורי דרך")
        st.markdown("שמור כ-PWA בסאפרי לגישה מהירה מהמסך הראשי")

    data = load_data()

    st.markdown("<h1 style='text-align:center; color:#1e293b;'>🥗 מעקב קלוריות ויעדי הרזיה</h1>",
                unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["📋 פרופיל אישי", "📝 יומן יומי", "📊 דשבורד"])

    with tab1:
        page_profile(data)
    with tab2:
        page_diary(data)
    with tab3:
        page_dashboard(data)


if __name__ == "__main__":
    main()
